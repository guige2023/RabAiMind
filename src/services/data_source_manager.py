# -*- coding: utf-8 -*-
"""
Data Source Manager Service (R75)
管理数据源：Excel/CSV 文件和 Google Sheets
支持：导入、同步、双向更新、自动刷新
"""

import io
import json
import logging
import os
import time
import uuid
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple

import httpx
from openpyxl import load_workbook

# Use shared HTTP client with connection pooling
from src.core.http_client import http_client

logger = logging.getLogger(__name__)

# 全局数据源存储（生产环境应使用数据库）
_data_sources: Dict[str, Dict[str, Any]] = {}
_storage_path = os.path.join(os.path.dirname(__file__), "..", "data", "data_sources.json")


def _load_storage() -> None:
    """从磁盘加载数据源存储"""
    global _data_sources
    try:
        if os.path.exists(_storage_path):
            with open(_storage_path, "r", encoding="utf-8") as f:
                _data_sources = json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load data sources storage: {e}")
        _data_sources = {}


def _save_storage() -> None:
    """保存数据源到磁盘"""
    try:
        os.makedirs(os.path.dirname(_storage_path), exist_ok=True)
        with open(_storage_path, "w", encoding="utf-8") as f:
            json.dump(_data_sources, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Failed to save data sources storage: {e}")


_load_storage()


class DataSourceManager:
    """数据源管理器"""

    # Google Sheets API base
    SHEETS_API = "https://sheets.googleapis.com/v4/spreadsheets"

    # 支持的文件扩展名
    SUPPORTED_EXCEL = {".xlsx", ".xls"}
    SUPPORTED_CSV = {".csv", ".tsv"}

    # Excel 表转幻灯片配置
    TABLE_LAYOUT_HINTS = {
        "sales": ["销售额", "收入", "sales", "revenue"],
        "user": ["用户", "新增用户", "活跃用户", "users"],
        "performance": ["性能", "响应时间", "QPS", "performance"],
    }

    def __init__(self):
        pass

    # ==================== 通用 ====================

    def list_data_sources(self, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """列出所有数据源"""
        _load_storage()
        sources = list(_data_sources.values())
        if user_id:
            sources = [s for s in sources if s.get("user_id") == user_id]
        return sorted(sources, key=lambda x: x.get("created_at", ""), reverse=True)

    def get_data_source(self, source_id: str) -> Optional[Dict[str, Any]]:
        """获取单个数据源"""
        _load_storage()
        return _data_sources.get(source_id)

    def create_data_source(
        self,
        name: str,
        source_type: str,
        user_id: Optional[str] = None,
        file_path: Optional[str] = None,
        file_name: Optional[str] = None,
        spreadsheet_url: Optional[str] = None,
        spreadsheet_id: Optional[str] = None,
        sheet_name: Optional[str] = None,
        extracted_data: Optional[Dict[str, Any]] = None,
        auto_update: bool = False,
    ) -> Dict[str, Any]:
        """创建新数据源"""
        source_id = str(uuid.uuid4())[:16]
        now = datetime.now().isoformat()

        source = {
            "id": source_id,
            "name": name,
            "source_type": source_type,
            "status": "active",
            "file_path": file_path,
            "file_name": file_name,
            "spreadsheet_url": spreadsheet_url,
            "spreadsheet_id": spreadsheet_id,
            "sheet_name": sheet_name,
            "auto_update": auto_update,
            "last_synced_at": now,
            "sync_interval_minutes": 60,
            "extracted_data": extracted_data,
            "created_at": now,
            "updated_at": now,
            "user_id": user_id,
        }

        _data_sources[source_id] = source
        _save_storage()
        logger.info(f"Created data source: {source_id} ({name})")
        return source

    def update_data_source(
        self,
        source_id: str,
        name: Optional[str] = None,
        auto_update: Optional[bool] = None,
        sync_interval_minutes: Optional[int] = None,
        status: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """更新数据源"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            return None

        if name is not None:
            source["name"] = name
        if auto_update is not None:
            source["auto_update"] = auto_update
        if sync_interval_minutes is not None:
            source["sync_interval_minutes"] = sync_interval_minutes
        if status is not None:
            source["status"] = status

        source["updated_at"] = datetime.now().isoformat()
        _data_sources[source_id] = source
        _save_storage()
        return source

    def delete_data_source(self, source_id: str) -> bool:
        """删除数据源"""
        _load_storage()
        if source_id in _data_sources:
            source = _data_sources.pop(source_id)
            _save_storage()
            # 删除关联文件
            if source.get("file_path") and os.path.exists(source["file_path"]):
                try:
                    os.remove(source["file_path"])
                except Exception:
                    pass
            logger.info(f"Deleted data source: {source_id}")
            return True
        return False

    # ==================== Excel/CSV 导入 ====================

    def import_excel_file(
        self,
        file_content: bytes,
        filename: str,
        sheet_index: int = 0,
        has_header: bool = True,
        max_rows: int = 1000,
        user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """导入 Excel 文件"""
        ext = os.path.splitext(filename)[-1].lower()
        if ext not in self.SUPPORTED_EXCEL:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx 和 .xls")

        try:
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Excel 文件解析失败: {e}")

        sheet_names = wb.sheetnames
        if sheet_index >= len(sheet_names):
            raise ValueError(f"工作表索引 {sheet_index} 超出范围，共 {len(sheet_names)} 个工作表")

        sheet = wb[sheet_names[sheet_index]]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError("Excel 文件为空")

        # 解析数据
        headers = []
        data_rows = []

        if has_header and len(rows) > 0:
            headers = [str(h) if h is not None else f"列{i}" for i, h in enumerate(rows[0])]
            data_rows = [list(row) for row in rows[1 : max_rows + 1]]
        else:
            headers = [f"列{i}" for i in range(len(rows[0]))]
            data_rows = [list(row) for row in rows[:max_rows]]

        # 构建表格预览 (前20行)
        preview = [headers] + data_rows[:20]

        # 列信息
        column_info = []
        for i, h in enumerate(headers):
            col_values = [row[i] if i < len(row) else None for row in data_rows[:100]]
            numeric_count = sum(1 for v in col_values if isinstance(v, (int, float)) and v is not None)
            col_type = "number" if numeric_count > len(col_values) * 0.5 else "text"
            column_info.append({
                "index": i,
                "name": h,
                "type": col_type,
                "sample": col_values[0] if col_values else None,
            })

        # 智能检测表格类型
        table_type = self._detect_table_type(headers, data_rows)

        # 转换为幻灯片大纲
        outline = self._table_to_slides(headers, data_rows, table_type)

        extracted_data = {
            "headers": headers,
            "data_rows": data_rows,
            "total_rows": len(data_rows),
            "total_columns": len(headers),
            "sheet_names": sheet_names,
            "current_sheet": sheet_names[sheet_index],
            "table_type": table_type,
            "column_info": column_info,
        }

        # 保存文件到临时目录
        upload_dir = os.path.join(os.path.dirname(__file__), "..", "data", "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        saved_path = os.path.join(upload_dir, f"{uuid.uuid4().hex[:8]}_{filename}")
        with open(saved_path, "wb") as f:
            f.write(file_content)

        # 创建数据源
        source = self.create_data_source(
            name=filename,
            source_type="excel",
            user_id=user_id,
            file_path=saved_path,
            file_name=filename,
            extracted_data=extracted_data,
            auto_update=False,
        )

        return {
            "success": True,
            "source_id": source["id"],
            "source_name": source["name"],
            "sheet_names": sheet_names,
            "current_sheet": sheet_names[sheet_index],
            "total_rows": len(data_rows),
            "total_columns": len(headers),
            "headers": headers,
            "column_info": column_info,
            "table_type": table_type,
            "table_preview": preview,
            "outline": outline,
        }

    def import_csv_file(
        self,
        file_content: bytes,
        filename: str,
        has_header: bool = True,
        max_rows: int = 1000,
        delimiter: str = ",",
        user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """导入 CSV 文件"""
        ext = os.path.splitext(filename)[-1].lower()
        if ext not in self.SUPPORTED_CSV:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .csv")

        try:
            # 自动检测分隔符
            text = file_content.decode("utf-8", errors="replace")
            if delimiter == "auto":
                comma_count = text.count(",")
                tab_count = text.count("\t")
                semicolon_count = text.count(";")
                if tab_count > comma_count and tab_count > semicolon_count:
                    delimiter = "\t"
                elif semicolon_count > comma_count:
                    delimiter = ";"
                else:
                    delimiter = ","

            import csv as csv_lib
            reader = csv_lib.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)
        except Exception as e:
            raise ValueError(f"CSV 文件解析失败: {e}")

        if not rows:
            raise ValueError("CSV 文件为空")

        headers = []
        data_rows = []

        if has_header and len(rows) > 0:
            headers = [str(h).strip() if h else f"列{i}" for i, h in enumerate(rows[0])]
            data_rows = [row for row in rows[1 : max_rows + 1]]
        else:
            headers = [f"列{i}" for i in range(len(rows[0]))]
            data_rows = rows[:max_rows]

        # 补齐列数不一致的行
        max_cols = len(headers)
        data_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]

        preview = [headers] + data_rows[:20]

        column_info = []
        for i, h in enumerate(headers):
            col_values = [row[i] if i < len(row) else None for row in data_rows[:100]]
            try:
                float(col_values[0])
                col_type = "number"
            except (TypeError, ValueError):
                col_type = "text"
            column_info.append({
                "index": i,
                "name": h,
                "type": col_type,
                "sample": col_values[0] if col_values else None,
            })

        table_type = self._detect_table_type(headers, data_rows)
        outline = self._table_to_slides(headers, data_rows, table_type)

        extracted_data = {
            "headers": headers,
            "data_rows": data_rows,
            "total_rows": len(data_rows),
            "total_columns": len(headers),
            "delimiter": delimiter,
            "table_type": table_type,
            "column_info": column_info,
        }

        upload_dir = os.path.join(os.path.dirname(__file__), "..", "data", "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        saved_path = os.path.join(upload_dir, f"{uuid.uuid4().hex[:8]}_{filename}")
        with open(saved_path, "wb") as f:
            f.write(file_content)

        source = self.create_data_source(
            name=filename,
            source_type="csv",
            user_id=user_id,
            file_path=saved_path,
            file_name=filename,
            extracted_data=extracted_data,
            auto_update=False,
        )

        return {
            "success": True,
            "source_id": source["id"],
            "source_name": source["name"],
            "total_rows": len(data_rows),
            "total_columns": len(headers),
            "headers": headers,
            "column_info": column_info,
            "table_type": table_type,
            "table_preview": preview,
            "outline": outline,
        }

    def _detect_table_type(self, headers: List[str], data_rows: List[List]) -> str:
        """智能检测表格类型"""
        header_text = " ".join(headers).lower()

        type_keywords = {
            "sales": ["销售额", "收入", "销量", "sales", "revenue", "订单", "成交"],
            "user": ["用户", "会员", "新增", "活跃", "users", "customers"],
            "marketing": ["营销", "推广", "广告", "投放", "marketing", "campaign"],
            "financial": ["成本", "利润", "支出", "预算", "financial", "cost", "profit"],
            "performance": ["性能", "响应", "QPS", "PV", "UV", "performance", "latency"],
            "inventory": ["库存", "商品", "SKU", "inventory", "stock"],
            "survey": ["满意度", "评分", "评价", "survey", "rating", "feedback"],
        }

        for table_type, keywords in type_keywords.items():
            if any(kw in header_text for kw in keywords):
                return table_type

        # 根据数值列比例判断
        numeric_cols = 0
        for i in range(min(len(headers), 10)):
            col_vals = [row[i] if i < len(row) else None for row in data_rows[:50]]
            if sum(1 for v in col_vals if isinstance(v, (int, float))) > len(col_vals) * 0.5:
                numeric_cols += 1

        if numeric_cols >= 3:
            return "data"
        return "general"

    def _table_to_slides(
        self,
        headers: List[str],
        data_rows: List[List],
        table_type: str,
        max_slides: int = 10,
    ) -> Dict[str, Any]:
        """将表格数据转换为 PPT 大纲"""
        total_rows = len(data_rows)
        rows_per_slide = max(1, total_rows // max_slides)

        slides = []

        # 封面
        slides.append({
            "slide_type": "title",
            "title": f"{headers[0] if headers else '数据报告'}",
            "subtitle": f"共 {total_rows} 条数据 · {len(headers)} 个指标",
        })

        # 数据概览页
        slides.append({
            "slide_type": "overview",
            "title": "数据概览",
            "content": self._generate_overview_text(headers, data_rows, table_type),
        })

        # 分组数据页
        for group_idx in range(max_slides):
            start = group_idx * rows_per_slide
            end = min(start + rows_per_slide, total_rows)
            if start >= total_rows:
                break

            group_data = data_rows[start:end]
            slides.append({
                "slide_type": "data_table",
                "title": f"{headers[0] if headers else '数据'} ({start + 1}-{end})",
                "headers": headers[:8],
                "rows": [row[:8] for row in group_data],
                "column_info": None,
            })

        # 汇总页
        slides.append({
            "slide_type": "summary",
            "title": "关键发现",
            "content": self._generate_summary_text(headers, data_rows, table_type),
        })

        return {
            "title": headers[0] if headers else "数据报告",
            "slides": slides,
        }

    def _generate_overview_text(
        self,
        headers: List[str],
        data_rows: List[List],
        table_type: str,
    ) -> str:
        """生成数据概览文本"""
        lines = [f"- 数据指标：{', '.join(headers[:5])}"]
        if len(headers) > 5:
            lines[0] += f" 等共 {len(headers)} 项"
        lines.append(f"- 数据记录：{len(data_rows)} 条")

        # 数值统计
        for i in range(min(3, len(headers))):
            col_vals = []
            for row in data_rows[:100]:
                if i < len(row):
                    try:
                        col_vals.append(float(row[i]))
                    except (TypeError, ValueError):
                        pass

            if col_vals:
                lines.append(
                    f"- {headers[i]}：总和 {sum(col_vals):.1f}，"
                    f"平均 {sum(col_vals)/len(col_vals):.1f}，"
                    f"最大 {max(col_vals):.1f}"
                )

        return "\n".join(lines)

    def _generate_summary_text(
        self,
        headers: List[str],
        data_rows: List[List],
        table_type: str,
    ) -> str:
        """生成关键发现文本"""
        lines = []

        if len(data_rows) > 0:
            lines.append(f"✓ 共分析了 {len(data_rows)} 条记录")

        # 找出数值变化最大的列
        max_change_col = -1
        max_change_val = 0
        for col_idx in range(1, min(len(headers), 6)):
            col_vals = []
            for row in data_rows:
                if col_idx < len(row):
                    try:
                        col_vals.append(float(row[col_idx]))
                    except (TypeError, ValueError):
                        pass

            if len(col_vals) >= 2:
                half = len(col_vals) // 2
                first_half_avg = sum(col_vals[:half]) / half if half > 0 else 0
                second_half_avg = sum(col_vals[half:]) / (len(col_vals) - half) if half > 0 else 0
                if first_half_avg > 0:
                    change = abs(second_half_avg - first_half_avg) / first_half_avg
                    if change > max_change_val:
                        max_change_val = change
                        max_change_col = col_idx

        if max_change_col >= 0 and max_change_val > 0.1:
            direction = "上升" if data_rows[-1][max_change_col] > data_rows[0][max_change_col] else "下降"
            lines.append(f"✓ {headers[max_change_col]} 整体呈{direction}趋势")
            lines.append(f"  变化幅度：{max_change_val*100:.1f}%")

        if len(data_rows) > 0 and len(headers) > 1:
            max_row_idx = max(range(len(data_rows)), key=lambda i: float(data_rows[i][1]) if len(data_rows[i]) > 1 else 0)
            if len(data_rows[max_row_idx]) > 0:
                lines.append(f"✓ 最大值出现在：{data_rows[max_row_idx][0]}")

        return "\n".join(lines) if lines else "✓ 数据分析完成"

    # ==================== Google Sheets 导入 ====================

    async def import_google_sheets(
        self,
        spreadsheet_url: str,
        sheet_name: Optional[str] = None,
        access_token: Optional[str] = None,
        user_id: Optional[str] = None,
        has_header: bool = True,
        max_rows: int = 1000,
    ) -> Dict[str, Any]:
        """从 Google Sheets 导入数据"""
        # 解析 spreadsheet ID
        spreadsheet_id = self._parse_spreadsheet_id(spreadsheet_url)
        if not spreadsheet_id:
            raise ValueError(f"无效的 Google Sheets URL: {spreadsheet_url}")

        if not access_token:
            raise ValueError("需要 Google OAuth access_token 来访问 Google Sheets")

        req_headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        }

        try:
            # 获取工作表信息（使用共享连接池）
            meta_resp = await http_client.get(
                f"{self.SHEETS_API}/{spreadsheet_id}",
                headers=req_headers,
                timeout=httpx.Timeout(30.0)
            )

            if meta_resp.status_code == 401:
                raise ValueError("Google Sheets 访问令牌已过期，请重新授权")
            if meta_resp.status_code == 403:
                raise ValueError("无权访问此 Google Sheets，请检查权限设置")
            if meta_resp.status_code != 200:
                raise ValueError(f"Google Sheets API 错误: {meta_resp.status_code}")

            meta = meta_resp.json()
            sheet_title = meta.get("properties", {}).get("title", "Sheet1")

            # 确定要读取的工作表
            target_sheet = sheet_name or sheet_title
            range_str = f"{target_sheet}!A:Z"

            # 读取数据（使用共享连接池）
            data_resp = await http_client.get(
                f"{self.SHEETS_API}/{spreadsheet_id}/values/{range_str}",
                headers=req_headers,
                params={"valueRenderOption": "FORMATTED_VALUE"},
                timeout=httpx.Timeout(30.0)
            )

            if data_resp.status_code != 200:
                raise ValueError(f"读取 Google Sheets 数据失败: {data_resp.status_code}")

                values = data_resp.json().get("values", [])

        except httpx.HTTPError as e:
            raise ValueError(f"Google Sheets 连接失败: {e}")

        if not values:
            raise ValueError("Google Sheets 为空")

        # 解析数据
        if has_header and len(values) > 0:
            sheet_headers = [str(h) if h else f"列{i}" for i, h in enumerate(values[0])]
            data_rows = [row for row in values[1 : max_rows + 1]]
        else:
            sheet_headers = [f"列{i}" for i in range(len(values[0]))]
            data_rows = values[:max_rows]

        # 补齐列
        max_cols = len(sheet_headers)
        data_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]

        preview = [sheet_headers] + data_rows[:20]

        column_info = []
        for i, h in enumerate(sheet_headers):
            col_vals = [row[i] if i < len(row) else None for row in data_rows[:100]]
            try:
                float(col_vals[0])
                col_type = "number"
            except (TypeError, ValueError):
                col_type = "text"
            column_info.append({
                "index": i,
                "name": h,
                "type": col_type,
                "sample": col_vals[0] if col_vals else None,
            })

        table_type = self._detect_table_type(sheet_headers, data_rows)
        outline = self._table_to_slides(sheet_headers, data_rows, table_type)

        extracted_data = {
            "headers": sheet_headers,
            "data_rows": data_rows,
            "total_rows": len(data_rows),
            "total_columns": len(sheet_headers),
            "spreadsheet_title": sheet_title,
            "table_type": table_type,
            "column_info": column_info,
        }

        source = self.create_data_source(
            name=f"Google Sheets: {sheet_title}",
            source_type="google_sheets",
            user_id=user_id,
            spreadsheet_url=spreadsheet_url,
            spreadsheet_id=spreadsheet_id,
            sheet_name=target_sheet,
            extracted_data=extracted_data,
            auto_update=False,
        )

        return {
            "success": True,
            "source_id": source["id"],
            "source_name": source["name"],
            "spreadsheet_title": sheet_title,
            "total_rows": len(data_rows),
            "total_columns": len(sheet_headers),
            "headers": sheet_headers,
            "column_info": column_info,
            "table_type": table_type,
            "table_preview": preview,
            "outline": outline,
        }

    def _parse_spreadsheet_id(self, url: str) -> Optional[str]:
        """从 URL 解析 spreadsheet ID"""
        import re
        patterns = [
            r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
            r"docs.google.com/spreadsheets/d/([a-zA-Z0-9-_]+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None

    # ==================== 同步 ====================

    async def sync_google_sheet(
        self,
        source_id: str,
        access_token: str,
    ) -> Dict[str, Any]:
        """同步 Google Sheets 数据"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            raise ValueError(f"数据源 {source_id} 不存在")

        if source["source_type"] != "google_sheets":
            raise ValueError("只有 Google Sheets 数据源支持同步")

        source["status"] = "syncing"
        _data_sources[source_id] = source
        _save_storage()

        try:
            spreadsheet_id = source.get("spreadsheet_id")
            sheet_name = source.get("sheet_name")

            if not access_token:
                raise ValueError("缺少 access_token，需要重新授权 Google Sheets")

            req_headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            }

            target_sheet = sheet_name or "Sheet1"
            range_str = f"{target_sheet}!A:Z"

            # 使用共享连接池
            data_resp = await http_client.get(
                f"{self.SHEETS_API}/{spreadsheet_id}/values/{range_str}",
                headers=req_headers,
                params={"valueRenderOption": "FORMATTED_VALUE"},
                timeout=httpx.Timeout(30.0)
            )

            if data_resp.status_code != 200:
                raise ValueError(f"同步失败: {data_resp.status_code}")

            values = data_resp.json().get("values", [])
            if not values:
                raise ValueError("Google Sheets 数据为空")

            has_header = True
            if has_header and len(values) > 0:
                sheet_headers = [str(h) if h else f"列{i}" for i, h in enumerate(values[0])]
                data_rows = values[1:]
            else:
                sheet_headers = [f"列{i}" for i in range(len(values[0]))]
                data_rows = values

            max_cols = len(sheet_headers)
            data_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]

            table_type = self._detect_table_type(sheet_headers, data_rows)
            outline = self._table_to_slides(sheet_headers, data_rows, table_type)

            now = datetime.now().isoformat()
            source["status"] = "active"
            source["last_synced_at"] = now
            source["updated_at"] = now
            source["extracted_data"] = {
                "headers": sheet_headers,
                "data_rows": data_rows,
                "total_rows": len(data_rows),
                "total_columns": len(sheet_headers),
                "table_type": table_type,
            }
            _data_sources[source_id] = source
            _save_storage()

            return {
                "success": True,
                "data_source_id": source_id,
                "synced_rows": len(data_rows),
                "synced_at": now,
                "message": "同步成功",
                "outline": outline,
            }

        except Exception as e:
            source = _data_sources.get(source_id)
            if source:
                source["status"] = "error"
                _data_sources[source_id] = source
                _save_storage()
            raise ValueError(f"同步失败: {e}")


# 单例
_manager = None


def get_data_source_manager() -> DataSourceManager:
    global _manager
    if _manager is None:
        _manager = DataSourceManager()
    return _manager


    # ==================== 阈值告警 (R113) ====================

    def check_threshold_alerts(
        self,
        source_id: str,
        alerts: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """检查阈值告警，返回触发的告警列表"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            return []

        ed = source.get("extracted_data", {})
        data_rows = ed.get("data_rows", [])
        headers = ed.get("headers", [])

        triggered = []
        for alert in alerts:
            if not alert.get("enabled", True):
                continue

            col_name = alert.get("column")
            condition = alert.get("condition", "gt")
            threshold = float(alert.get("value", 0))

            col_idx = -1
            for i, h in enumerate(headers):
                if h == col_name:
                    col_idx = i
                    break

            if col_idx < 0:
                continue

            for row_idx, row in enumerate(data_rows):
                if col_idx >= len(row):
                    continue
                try:
                    cell_val = float(row[col_idx])
                except (TypeError, ValueError):
                    continue

                is_triggered = False
                if condition == "gt" and cell_val > threshold:
                    is_triggered = True
                elif condition == "lt" and cell_val < threshold:
                    is_triggered = True
                elif condition == "gte" and cell_val >= threshold:
                    is_triggered = True
                elif condition == "lte" and cell_val <= threshold:
                    is_triggered = True
                elif condition == "eq" and abs(cell_val - threshold) < 0.0001:
                    is_triggered = True

                if is_triggered:
                    triggered.append({
                        "alert_label": alert.get("label", col_name),
                        "column": col_name,
                        "condition": condition,
                        "threshold": threshold,
                        "actual_value": cell_val,
                        "row_label": str(row[0]) if row else f"Row {row_idx}",
                        "row_index": row_idx,
                        "highlight": True,
                    })

        return triggered

    def update_threshold_alerts(self, source_id: str, alerts: List[Dict[str, Any]]) -> None:
        """更新阈值告警规则"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            return

        source["threshold_alerts"] = alerts
        source["updated_at"] = datetime.now().isoformat()
        _data_sources[source_id] = source
        _save_storage()

    def analyze_data(
        self,
        source_id: str,
        compare_column: str,
        group_by_column: Optional[str] = None,
        periods: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """分析数据：聚合、对比、趋势"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            raise ValueError(f"数据源 {source_id} 不存在")

        ed = source.get("extracted_data", {})
        data_rows = ed.get("data_rows", [])
        headers = ed.get("headers", [])

        compare_idx = -1
        for i, h in enumerate(headers):
            if h == compare_column:
                compare_idx = i
                break
        if compare_idx < 0:
            raise ValueError(f"未找到列: {compare_column}")

        group_idx = -1
        if group_by_column:
            for i, h in enumerate(headers):
                if h == group_by_column:
                    group_idx = i
                    break

        compare_values = []
        group_labels = []
        row_labels = []

        for row_idx, row in enumerate(data_rows):
            if compare_idx >= len(row):
                continue
            try:
                val = float(row[compare_idx])
                compare_values.append(val)
                if group_idx >= 0 and group_idx < len(row):
                    group_labels.append(str(row[group_idx]))
                else:
                    group_labels.append(f"Item {row_idx + 1}")
                row_labels.append(str(row[0]) if row else f"Row {row_idx + 1}")
            except (TypeError, ValueError):
                continue

        if not compare_values:
            raise ValueError(f"列 {compare_column} 没有数值数据")

        total = sum(compare_values)
        avg = total / len(compare_values)
        max_val = max(compare_values)
        min_val = min(compare_values)
        max_idx = compare_values.index(max_val)
        min_idx = compare_values.index(min_val)

        half = len(compare_values) // 2
        first_half_avg = sum(compare_values[:half]) / half if half > 0 else 0
        second_half_avg = sum(compare_values[half:]) / (len(compare_values) - half) if half > 0 else 0

        trend = "stable"
        trend_pct = 0.0
        if first_half_avg > 0:
            trend_pct = (second_half_avg - first_half_avg) / first_half_avg * 100
            if trend_pct > 5:
                trend = "increasing"
            elif trend_pct < -5:
                trend = "decreasing"

        groups = {}
        if group_idx >= 0:
            for i, (val, g_label) in enumerate(zip(compare_values, group_labels)):
                if g_label not in groups:
                    groups[g_label] = []
                groups[g_label].append(val)

        group_stats = {}
        for g_label, vals in groups.items():
            if vals:
                group_stats[g_label] = {
                    "sum": round(sum(vals), 2),
                    "avg": round(sum(vals) / len(vals), 2),
                    "count": len(vals),
                    "max": round(max(vals), 2),
                    "min": round(min(vals), 2),
                }

        comparison = {}
        if periods and len(periods) >= 2:
            period_len = len(data_rows) // len(periods)
            for i, period in enumerate(periods):
                start = i * period_len
                end = start + period_len
                vals = compare_values[start:end]
                if vals:
                    comparison[period] = {
                        "sum": round(sum(vals), 2),
                        "avg": round(sum(vals) / len(vals), 2),
                    }

        return {
            "success": True,
            "source_id": source_id,
            "compare_column": compare_column,
            "total_rows": len(compare_values),
            "stats": {
                "sum": round(total, 2),
                "avg": round(avg, 2),
                "max": round(max_val, 2),
                "min": round(min_val, 2),
                "max_label": row_labels[max_idx],
                "min_label": row_labels[min_idx],
            },
            "trend": {
                "direction": trend,
                "change_pct": round(trend_pct, 2),
                "first_half_avg": round(first_half_avg, 2),
                "second_half_avg": round(second_half_avg, 2),
            },
            "group_stats": group_stats if group_stats else None,
            "period_comparison": comparison if comparison else None,
            "chart_data": {
                "labels": row_labels[:20],
                "values": [round(v, 2) for v in compare_values[:20]],
            },
        }

    def generate_forecast(
        self,
        source_id: str,
        value_column: str,
        label_column: Optional[str] = None,
        forecast_periods: int = 3,
        chart_type: str = "line",
    ) -> Dict[str, Any]:
        """生成预测图表数据（线性回归趋势 + 未来预测）"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            raise ValueError(f"数据源 {source_id} 不存在")

        ed = source.get("extracted_data", {})
        data_rows = ed.get("data_rows", [])
        headers = ed.get("headers", [])

        value_idx = -1
        for i, h in enumerate(headers):
            if h == value_column:
                value_idx = i
                break
        if value_idx < 0:
            raise ValueError(f"未找到列: {value_column}")

        label_idx = 0
        if label_column:
            for i, h in enumerate(headers):
                if h == label_column:
                    label_idx = i
                    break

        values = []
        labels = []

        for row_idx, row in enumerate(data_rows):
            if value_idx >= len(row):
                continue
            try:
                val = float(row[value_idx])
                values.append(val)
                if label_idx < len(row):
                    labels.append(str(row[label_idx]))
                else:
                    labels.append(f"T{row_idx + 1}")
            except (TypeError, ValueError):
                continue

        if len(values) < 3:
            raise ValueError("数据行数不足3行，无法进行预测")

        n = len(values)
        x_vals = list(range(n))
        y_vals = values

        x_mean = sum(x_vals) / n
        y_mean = sum(y_vals) / n

        numerator = sum((x_vals[i] - x_mean) * (y_vals[i] - y_mean) for i in range(n))
        denominator = sum((x_vals[i] - x_mean) ** 2 for i in range(n))

        if denominator == 0:
            slope = 0
            intercept = y_mean
        else:
            slope = numerator / denominator
            intercept = y_mean - slope * x_mean

        trendline = [round(slope * x + intercept, 2) for x in x_vals]

        ss_res = sum((y_vals[i] - trendline[i]) ** 2 for i in range(n))
        ss_tot = sum((y_vals[i] - y_mean) ** 2 for i in range(n))
        r_squared = round(1 - ss_res / ss_tot, 4) if ss_tot > 0 else 0

        forecast_labels = []
        forecast_values = []
        last_label = labels[-1] if labels else f"T{n}"

        for i in range(1, forecast_periods + 1):
            future_x = n + i - 1
            future_val = round(slope * future_x + intercept, 2)
            forecast_values.append(future_val)
            if last_label and last_label[0].isdigit():
                import re
                nums = re.findall(r'\d+', last_label)
                base_num = int(nums[-1]) if nums else i
                forecast_labels.append(str(base_num + i))
            else:
                forecast_labels.append(f"{last_label}+{i}")

        chart_labels = labels[:n] + forecast_labels
        trendline_extended = trendline + [round(slope * (n + i - 1) + intercept, 2) for i in range(1, forecast_periods + 1)]

        return {
            "success": True,
            "source_id": source_id,
            "value_column": value_column,
            "regression": {
                "slope": round(slope, 4),
                "intercept": round(intercept, 2),
                "r_squared": r_squared,
            },
            "trend_direction": "up" if slope > 0 else "down" if slope < 0 else "flat",
            "trendline": trendline,
            "forecast": {
                "periods": forecast_periods,
                "labels": forecast_labels,
                "values": forecast_values,
            },
            "chart_data": {
                "type": chart_type,
                "labels": chart_labels,
                "actual": [round(v, 2) for v in values] + [None] * forecast_periods,
                "forecast": [None] * n + forecast_values,
                "trendline": trendline_extended,
            },
        }

    def generate_ppt_from_data_source(
        self,
        source_id: str,
        title: Optional[str] = None,
        include_charts: bool = True,
        include_threshold_alerts: bool = True,
        include_forecast: bool = False,
        forecast_periods: int = 3,
        slide_count: int = 10,
        user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """使用数据源生成 PPT 大纲"""
        _load_storage()
        source = _data_sources.get(source_id)
        if not source:
            raise ValueError(f"数据源 {source_id} 不存在")

        ed = source.get("extracted_data", {})
        headers = ed.get("headers", [])
        data_rows = ed.get("data_rows", [])
        table_type = ed.get("table_type", "general")

        ppt_title = title or f"{source['name']} 数据报告"
        slides = []

        slides.append({
            "slide_type": "title",
            "title": ppt_title,
            "subtitle": f"数据来源: {source['name']} - {len(data_rows)} 条记录",
        })

        slides.append({
            "slide_type": "overview",
            "title": "数据概览",
            "content": self._generate_overview_text(headers, data_rows, table_type),
        })

        if include_charts and len(headers) >= 2:
            numeric_cols = []
            for i, col_info in enumerate(ed.get("column_info", [])):
                if col_info.get("type") == "number":
                    numeric_cols.append((i, col_info.get("name", headers[i])))

            if numeric_cols:
                slides.append({
                    "slide_type": "chart_description",
                    "title": "数据分析图表",
                    "charts": [
                        {
                            "chart_type": "bar",
                            "title": f"{numeric_cols[0][1]} 分布",
                            "column": numeric_cols[0][1],
                        }
                    ],
                })

        if include_threshold_alerts:
            alerts = source.get("threshold_alerts", [])
            if alerts:
                triggered = self.check_threshold_alerts(source_id, alerts)
                if triggered:
                    alert_text = "\n".join(
                        [f"!! {a['alert_label']}: {a['actual_value']} ({a['condition']} {a['threshold']})"
                         for a in triggered[:5]]
                    )
                    slides.append({
                        "slide_type": "alert",
                        "title": "!! 阈值告警",
                        "content": alert_text,
                        "triggered_count": len(triggered),
                    })

        if include_forecast and len(headers) >= 2:
            numeric_cols_for_forecast = []
            for i, col_info in enumerate(ed.get("column_info", [])):
                if col_info.get("type") == "number":
                    numeric_cols_for_forecast.append(col_info.get("name", headers[i]))

            if numeric_cols_for_forecast:
                try:
                    forecast_result = self.generate_forecast(
                        source_id=source_id,
                        value_column=numeric_cols_for_forecast[0],
                        forecast_periods=forecast_periods,
                        chart_type="line",
                    )
                    slides.append({
                        "slide_type": "forecast",
                        "title": f"{numeric_cols_for_forecast[0]} 趋势与预测",
                        "trend": forecast_result["trend_direction"],
                        "r_squared": forecast_result["regression"]["r_squared"],
                        "forecast_periods": forecast_periods,
                        "forecast_values": forecast_result["forecast"]["values"],
                        "trendline": forecast_result["trendline"][-5:] if forecast_result["trendline"] else [],
                    })
                except Exception:
                    pass

        rows_per_slide = max(1, len(data_rows) // (slide_count - 5))
        for group_idx in range(slide_count - 5):
            start = group_idx * rows_per_slide
            end = min(start + rows_per_slide, len(data_rows))
            if start >= len(data_rows):
                break

            group_data = data_rows[start:end]
            slides.append({
                "slide_type": "data_table",
                "title": f"数据 ({start + 1}-{end})",
                "headers": headers[:8],
                "rows": [row[:8] for row in group_data],
            })

        slides.append({
            "slide_type": "summary",
            "title": "关键发现",
            "content": self._generate_summary_text(headers, data_rows, table_type),
        })

        return {
            "success": True,
            "source_id": source_id,
            "outline": {
                "title": ppt_title,
                "slides": slides,
            },
        }
